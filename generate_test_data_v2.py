#!/usr/bin/env python3
"""
Generate test data for Fleet Management System
EXACT column names matching database schema
"""

import random
import uuid
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_FILE = "/tmp/fleet_test_data_v2.xlsx"

# Data pools
FIRST_NAMES = ["John", "Mary", "James", "Patricia", "Robert", "Jennifer", "Michael", "Linda", "William", "Elizabeth",
               "David", "Susan", "Richard", "Jessica", "Joseph", "Sarah", "Thomas", "Karen", "Charles", "Nancy"]

LAST_NAMES = ["Kamau", "Odhiambo", "Ochieng", "Wanjiku", "Kipchoge", "Mutua", "Omondi", "Wanjiru", "Mwangi", "Kiptoo"]

DEPARTMENTS = ["Transport", "Security", "Operations", "Finance", "HR", "IT", "Sales", "Logistics"]
BRANCHES = ["Nairobi HQ", "Mombasa", "Kisumu", "Nakuru", "Eldoret"]
ROLES = ["Driver", "Transport Supervisor", "Departmental Supervisor", "Head of Department", "Security Personnel"]
VEHICLE_MAKES = ["Toyota Hilux", "Toyota Land Cruiser", "Toyota Hiace", "Nissan Patrol", "Mitsubishi L200", "Ford Ranger"]
OWNERSHIP_TYPES = ["Company", "Leased", "Hired"]
REPAIR_TYPES = ["Oil Change", "Brake Replacement", "Tire Rotation", "Engine Tune-up", "Transmission Service"]
GARAGES = ["Toyota Kenya", "CMC Motors", "DT Dobie", "City Garage", "Highway Service Center"]
FUEL_STATIONS = ["Shell", "Total", "Oilibya", "National Oil", "Kobil"]
PLACES = ["Nairobi", "Mombasa", "Kisumu", "Nakuru", "Eldoret"]

def random_date(start_days=365, end_days=0):
    days = random.randint(end_days, start_days)
    return (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

def random_time():
    return f"{random.randint(6, 18):02d}:{random.randint(0, 59):02d}"

def generate_staff(count=20):
    staff = []
    for i in range(count):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        staff.append({
            "staff_no": f"ST{1001+i}",
            "staff_name": f"{first} {last}",
            "email": f"{first.lower()}.{last.lower()}@g4s.com",
            "phone": f"07{random.randint(10000000, 99999999)}",
            "designation": random.choice(["Senior Driver", "Driver", "Supervisor"]),
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "role": random.choice(ROLES),
            "comments": ""
        })
    return staff

def generate_vehicles(count=15):
    vehicles = []
    for i in range(count):
        ownership = random.choice(OWNERSHIP_TYPES)
        vehicles.append({
            "registration_num": f"K{random.choice(['A','B','C','X','Y'])}{random.randint(100,999)} {random.choice(['A','B','X'])}{random.randint(1000,9999)}",
            "year_of_manufacture": random.randint(2015, 2023),
            "year_of_purchase": random.randint(2016, 2024),
            "replacement_mileage": random.randint(150000, 300000),
            "replacement_age": random.randint(7, 12),
            "make_model": random.choice(VEHICLE_MAKES),
            "ownership": ownership,
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "minor_service_interval": 5000 if ownership == "Company" else 3000,
            "medium_service_interval": 15000 if ownership == "Company" else 10000,
            "major_service_interval": 30000 if ownership == "Company" else 20000,
            "target_consumption_rate": round(random.uniform(8.0, 15.0), 2),
            "status": random.choice(["Active", "Under Maintenance", "Retired"]),
            "current_mileage": random.randint(10000, 150000),
            "last_service_date": random_date(180, 30),
            "next_service_due": random_date(30, -30)
        })
    return vehicles

def generate_routes(staff, vehicles, count=30):
    """Generate routes with EXACT column names from database"""
    routes = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    
    for i in range(count):
        driver = random.choice(drivers)
        vehicle = random.choice(vehicles)
        target_km = random.randint(100, 800)
        actual_km = target_km + random.randint(-50, 100)
        target_fuel = round(target_km / random.uniform(8.0, 12.0), 2)
        actual_fuel = round(actual_km / random.uniform(7.5, 13.0), 2)
        
        routes.append({
            "route_date": random_date(60, 0),
            "route_name": f"Route {i+1}: {random.choice(PLACES)} to {random.choice(PLACES)}",
            "driver1_id": driver["staff_no"],  # Using staff_no for reference
            "driver2_id": "",
            "co_driver_id": "",
            "vehicle_id": vehicle["registration_num"],  # Using reg_num for reference
            "target_km": target_km,
            "actual_km": actual_km,
            "target_fuel_consumption": target_fuel,
            "actual_fuel": actual_fuel,
            "target_consumption_rate": round(random.uniform(8.0, 12.0), 2),
            "actual_consumption_rate": round(actual_fuel / actual_km * 100, 2) if actual_km > 0 else 0,
            "variance": round(actual_fuel - target_fuel, 2),
            "comments": ""
        })
    return routes

def generate_fuel(vehicles, count=50):
    fuel = []
    for i in range(count):
        vehicle = random.choice(vehicles)
        past = vehicle["current_mileage"] - random.randint(500, 3000)
        current = past + random.randint(400, 2500)
        qty = round((current - past) / random.uniform(8.0, 12.0), 2)
        amt = round(qty * random.uniform(170, 195), 2)
        
        fuel.append({
            "department": vehicle["department"],
            "fuel_date": random_date(90, 0),
            "vehicle_id": vehicle["registration_num"],
            "card_num": f"FC{random.randint(100000, 999999)}",
            "card_name": random.choice(FUEL_STATIONS),
            "past_mileage": past,
            "current_mileage": current,
            "quantity_liters": qty,
            "amount": amt,
            "place": random.choice(PLACES)
        })
    return fuel

def generate_repairs(vehicles, staff, count=25):
    repairs = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    
    for i in range(count):
        vehicle = random.choice(vehicles)
        driver = random.choice(drivers)
        date_in = random_date(120, 0)
        target = random.uniform(2, 24)
        actual = target + random.uniform(-2, 8)
        
        repairs.append({
            "date_in": date_in,
            "vehicle_id": vehicle["registration_num"],
            "preventative_maintenance": random.choice(REPAIR_TYPES),
            "breakdown_description": "Routine maintenance" if random.random() > 0.3 else "Minor repair",
            "odometer_reading": vehicle["current_mileage"],
            "driver_id": driver["staff_name"],
            "assigned_technician": random.choice(["John Mechanic", "Peter Fixer", "James Repair"]),
            "date_out": (datetime.strptime(date_in, '%Y-%m-%d') + timedelta(days=random.randint(1,5))).strftime('%Y-%m-%d'),
            "actual_repair_hours": round(actual, 2),
            "target_repair_hours": round(target, 2),
            "productivity_ratio": round(target / actual, 2) if actual > 0 else 1.0,
            "garage_name": random.choice(GARAGES),
            "cost": round(random.uniform(5000, 50000), 2),
            "status": random.choice(["Completed", "In Progress", "Pending"])
        })
    return repairs

def create_excel():
    print("🚀 Generating test data...")
    
    staff = generate_staff(20)
    vehicles = generate_vehicles(15)
    routes = generate_routes(staff, vehicles, 30)
    fuel = generate_fuel(vehicles, 50)
    repairs = generate_repairs(vehicles, staff, 25)
    
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 1. STAFF Sheet (exact column names)
    ws_staff = wb.active
    ws_staff.title = "Staff"
    staff_headers = ["staff_no", "staff_name", "email", "phone", "designation", "department", "branch", "role", "comments"]
    ws_staff.append(staff_headers)
    for cell in ws_staff[1]:
        cell.fill = header_fill
        cell.font = header_font
    for s in staff:
        ws_staff.append([s[h] for h in staff_headers])
    
    # 2. VEHICLES Sheet
    ws_veh = wb.create_sheet("Vehicles")
    veh_headers = ["registration_num", "year_of_manufacture", "year_of_purchase", "replacement_mileage",
                   "replacement_age", "make_model", "ownership", "department", "branch",
                   "minor_service_interval", "medium_service_interval", "major_service_interval",
                   "target_consumption_rate", "status", "current_mileage", "last_service_date", "next_service_due"]
    ws_veh.append(veh_headers)
    for cell in ws_veh[1]:
        cell.fill = header_fill
        cell.font = header_font
    for v in vehicles:
        ws_veh.append([v[h] for h in veh_headers])
    
    # 3. ROUTES Sheet (exact column names from database)
    ws_routes = wb.create_sheet("Routes")
    route_headers = ["route_date", "route_name", "driver1_id", "driver2_id", "co_driver_id", "vehicle_id",
                     "target_km", "actual_km", "target_fuel_consumption", "actual_fuel",
                     "target_consumption_rate", "actual_consumption_rate", "variance", "comments"]
    ws_routes.append(route_headers)
    for cell in ws_routes[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in routes:
        ws_routes.append([r[h] for h in route_headers])
    
    # 4. FUEL Sheet
    ws_fuel = wb.create_sheet("Fuel")
    fuel_headers = ["department", "fuel_date", "vehicle_id", "card_num", "card_name",
                    "past_mileage", "current_mileage", "quantity_liters", "amount", "place"]
    ws_fuel.append(fuel_headers)
    for cell in ws_fuel[1]:
        cell.fill = header_fill
        cell.font = header_font
    for f in fuel:
        ws_fuel.append([f[h] for h in fuel_headers])
    
    # 5. REPAIRS Sheet
    ws_rep = wb.create_sheet("Repairs")
    rep_headers = ["date_in", "vehicle_id", "preventative_maintenance", "breakdown_description",
                   "odometer_reading", "driver_id", "assigned_technician", "date_out",
                   "actual_repair_hours", "target_repair_hours", "productivity_ratio",
                   "garage_name", "cost", "status"]
    ws_rep.append(rep_headers)
    for cell in ws_rep[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in repairs:
        ws_rep.append([r[h] for h in rep_headers])
    
    # Auto-adjust columns
    for ws in [ws_staff, ws_veh, ws_routes, ws_fuel, ws_rep]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(OUTPUT_FILE)
    print(f"✅ Saved: {OUTPUT_FILE}")
    print(f"📊 Staff: {len(staff)}, Vehicles: {len(vehicles)}, Routes: {len(routes)}, Fuel: {len(fuel)}, Repairs: {len(repairs)}")
    return OUTPUT_FILE

if __name__ == "__main__":
    create_excel()
