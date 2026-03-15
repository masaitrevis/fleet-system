#!/usr/bin/env python3
"""
Fleet Management System - COMPLETE Test Data Generator
Columns EXACTLY match backend/upload.ts expectations
Drivers optimized for Analytics (routes + accidents linked)
"""

import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUTPUT_FILE = "/root/.openclaw/workspace/fleet-system/fleet_test_data_complete.xlsx"

# === DATA POOLS ===
FIRST_NAMES = ["John", "Mary", "James", "Patricia", "Robert", "Jennifer", "Michael", "Linda", 
               "William", "Elizabeth", "David", "Susan", "Richard", "Jessica", "Joseph", 
               "Sarah", "Thomas", "Karen", "Charles", "Nancy", "Daniel", "Lisa", "Matthew",
               "Betty", "Anthony", "Margaret", "Mark", "Sandra", "Donald", "Ashley",
               "Peter", "Nancy", "Kenneth", "Betty", "George", "Helen", "Paul", "Sandra"]

LAST_NAMES = ["Kamau", "Odhiambo", "Ochieng", "Wanjiku", "Kipchoge", "Mutua", "Omondi", 
              "Wanjiru", "Mwangi", "Kiptoo", "Korir", "Ruto", "Kiprotich", "Tanui",
              "Kiplagat", "Chebet", "Langat", "Kosgei", "Biwott", "Too", "Koech", "Kemei"]

DEPARTMENTS = ["Transport", "Security", "Operations", "Finance", "HR", "IT", "Sales", "Logistics"]
BRANCHES = ["Nairobi HQ", "Mombasa", "Kisumu", "Nakuru", "Eldoret", "Thika", "Malindi"]
DRIVER_NAMES = []  # Will be populated during staff generation

VEHICLE_MAKES = ["Toyota Hilux", "Toyota Land Cruiser 79", "Toyota Land Cruiser 200", "Toyota Hiace",
                 "Toyota Coaster", "Toyota Corolla", "Nissan Patrol", "Nissan NP300", 
                 "Mitsubishi L200", "Ford Ranger", "Isuzu D-Max", "Mazda BT-50"]
OWNERSHIP_TYPES = ["Company", "Leased", "Hired"]
VEHICLE_STATUS = ["Active", "Under Maintenance", "Retired"]

REPAIR_TYPES = ["Oil Change", "Brake Replacement", "Tire Rotation", "Engine Tune-up", 
                "Transmission Service", "Clutch Replacement", "Suspension Repair", 
                "Electrical Repair", "Air Conditioning", "Body Work"]
GARAGES = ["Toyota Kenya (Nairobi)", "Toyota Kenya (Mombasa)", "CMC Motors", "DT Dobie", 
           "City Garage", "Highway Service Center", "Kobil Workshop", "Shell Auto Care"]
FUEL_STATIONS = ["Shell", "Total", "Oilibya", "National Oil", "Kobil", "Rubis", "Delta"]

ROUTE_ORIGINS = ["Nairobi HQ", "JKIA", "Industrial Area", "Westlands", "Mombasa Road Depot",
                 "Mombasa Port", "Mombasa CBD", "Changamwe", "Kisumu Depot", "Kisumu Port",
                 "Nakuru Depot", "Eldoret Branch", "Thika Town"]
ROUTE_DESTINATIONS = ["Mombasa", "Nairobi", "Kisumu", "Nakuru", "Eldoret", "Malindi", 
                      "Thika", "Namanga Border", "Busia Border", "Taveta", "Nyeri", "Meru"]

ACCIDENT_TYPES = ["Collision", "Rollover", "Side-swipe", "Rear-end", "Pedestrian", "Animal", "Other"]
SEVERITY_LEVELS = ["Minor", "Moderate", "Major", "Fatal"]
WEATHER_CONDITIONS = ["Clear", "Rain", "Fog", "Night"]
ROAD_CONDITIONS = ["Dry", "Wet", "Muddy", "Under Construction"]

def random_date(start_days=365, end_days=0):
    days = random.randint(end_days, start_days)
    return (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

def random_datetime(start_days=365, end_days=0):
    days = random.randint(end_days, start_days)
    date = datetime.now() - timedelta(days=days)
    hour = random.randint(6, 22)
    minute = random.randint(0, 59)
    return date.replace(hour=hour, minute=minute).strftime('%Y-%m-%d %H:%M')

def generate_registration():
    prefixes = ['K', 'K', 'K', 'K', 'K', 'KB', 'KC', 'KX', 'KY']
    prefix = random.choice(prefixes)
    if len(prefix) == 1:
        letter = random.choice(['A', 'B', 'C', 'D', 'X', 'Y', 'Z'])
        numbers = random.randint(10, 999)
        suffix = random.choice(['A', 'B', 'C', 'X', 'Y'])
        suffix_num = random.randint(1000, 9999)
        return f"{prefix}{letter}{numbers} {suffix}{suffix_num}"
    else:
        numbers = random.randint(100, 999)
        suffix = random.choice(['A', 'B', 'C', 'X', 'Y'])
        suffix_num = random.randint(1000, 9999)
        return f"{prefix}{numbers} {suffix}{suffix_num}"

def generate_phone():
    prefixes = ['07', '01']
    return f"{random.choice(prefixes)}{random.randint(10000000, 99999999)}"

def generate_gps():
    # Kenya coordinates roughly
    lat = round(random.uniform(-4.6, 1.0), 6)
    lng = round(random.uniform(33.9, 41.9), 6)
    return f"{lat}, {lng}"

# === GENERATOR FUNCTIONS ===

def generate_staff(count=30):
    """Generate staff with proper roles for analytics"""
    staff = []
    global DRIVER_NAMES
    
    # Ensure we have drivers for analytics
    driver_count = max(10, count // 3)
    supervisor_count = max(5, count // 6)
    
    roles_list = (["Driver"] * driver_count + 
                  ["Transport Supervisor"] * supervisor_count +
                  ["Departmental Supervisor"] * 3 +
                  ["Head of Department"] * 2 +
                  ["Security Personnel"] * 3)
    
    # Pad to reach count
    while len(roles_list) < count:
        roles_list.append(random.choice(["Driver", "Transport Supervisor"]))
    roles_list = roles_list[:count]
    random.shuffle(roles_list)
    
    for i in range(count):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        role = roles_list[i]
        
        staff.append({
            "staff_no": f"ST{1001+i}",
            "staff_name": f"{first} {last}",
            "email": f"{first.lower()}.{last.lower()}@company.com",
            "phone": generate_phone(),
            "designation": "Senior Driver" if role == "Driver" and i % 3 == 0 else role,
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "role": role
        })
        
        if role == "Driver":
            DRIVER_NAMES.append(staff[-1]["staff_name"])
    
    return staff

def generate_vehicles(count=25):
    """Generate vehicles for the fleet"""
    vehicles = []
    for i in range(count):
        ownership = random.choice(OWNERSHIP_TYPES)
        year_mfg = random.randint(2015, 2023)
        current_mileage = random.randint(15000, 180000)
        
        vehicles.append({
            "registration_num": generate_registration(),
            "year_of_manufacture": year_mfg,
            "year_of_purchase": year_mfg + random.randint(0, 1),
            "make_model": random.choice(VEHICLE_MAKES),
            "ownership": ownership,
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "minor_service_interval": 5000,
            "medium_service_interval": 15000,
            "major_service_interval": 30000,
            "target_consumption_rate": round(random.uniform(8.0, 14.0), 2),
            "status": "Active" if random.random() > 0.15 else random.choice(["Under Maintenance", "Retired"]),
            "current_mileage": current_mileage
        })
    return vehicles

def generate_routes(staff, vehicles, count=60):
    """Generate routes assigned to drivers for analytics"""
    routes = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    active_vehicles = [v for v in vehicles if v["status"] == "Active"]
    
    if not drivers or not active_vehicles:
        return routes
    
    for i in range(count):
        driver = random.choice(drivers)
        vehicle = random.choice(active_vehicles)
        
        target_km = random.randint(100, 600)
        actual_km = max(50, target_km + random.randint(-30, 80))
        target_fuel = round(target_km / random.uniform(8.0, 12.0), 2)
        actual_fuel = round(actual_km / random.uniform(7.5, 13.0), 2)
        variance = round(actual_fuel - target_fuel, 2)
        
        origin = random.choice(ROUTE_ORIGINS)
        dest = random.choice(ROUTE_DESTINATIONS)
        while dest == origin:
            dest = random.choice(ROUTE_DESTINATIONS)
        
        routes.append({
            "route_date": random_date(90, 0),
            "route_name": f"{origin} - {dest}",
            "driver1": driver["staff_no"],  # Links to driver for analytics
            "vehicle": vehicle["registration_num"],
            "target_km": target_km,
            "actual_km": actual_km,
            "target_fuel": target_fuel,
            "actual_fuel": actual_fuel,
            "variance": variance
        })
    return routes

def generate_fuel(vehicles, count=100):
    """Generate fuel records linked to vehicles"""
    fuel = []
    active_vehicles = [v for v in vehicles if v["status"] == "Active"]
    
    for i in range(count):
        vehicle = random.choice(active_vehicles)
        current = random.randint(20000, 180000)
        past = current - random.randint(300, 2500)
        
        distance = current - past
        km_per_l = random.uniform(7.5, 13.0)
        qty = round(distance / km_per_l, 2)
        price_per_l = random.uniform(175, 198)
        amt = round(qty * price_per_l, 2)
        
        fuel.append({
            "department": vehicle["department"],
            "date": random_date(120, 0),
            "registration": vehicle["registration_num"],
            "card_num": f"FC{random.randint(100000, 999999)}",
            "card_name": random.choice(FUEL_STATIONS),
            "past": past,
            "current": current,
            "quantity": qty,
            "amount": amt,
            "place": random.choice(BRANCHES)
        })
    return fuel

def generate_repairs(vehicles, staff, count=40):
    """Generate repair records linked to vehicles"""
    repairs = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    
    for i in range(count):
        vehicle = random.choice(vehicles)
        driver = random.choice(drivers) if drivers and random.random() > 0.3 else None
        
        repairs.append({
            "date_in": random_date(180, 0),
            "registration": vehicle["registration_num"],
            "maintenance": random.choice(REPAIR_TYPES),
            "description": random.choice(["Routine service", "Minor repair", "Annual check", "Emergency repair"]),
            "odometer": random.randint(20000, 180000),
            "technician": random.choice(["John Mechanic", "Peter Fixer", "James Repair", "Sam Auto"]),
            "garage": random.choice(GARAGES),
            "assigned_driver": driver["staff_no"] if driver else ""
        })
    return repairs

def generate_accidents(staff, vehicles, count=15):
    """Generate accidents linked to drivers for analytics"""
    accidents = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    active_vehicles = [v for v in vehicles if v["status"] == "Active"]
    
    for i in range(count):
        driver = random.choice(drivers) if drivers else None
        vehicle = random.choice(active_vehicles) if active_vehicles else None
        severity = random.choice(SEVERITY_LEVELS)
        
        accidents.append({
            "case_number": f"ACC-2025-{str(i+1).zfill(4)}",
            "accident_date": random_datetime(365, 30),
            "gps_location": generate_gps(),
            "registration": vehicle["registration_num"] if vehicle else generate_registration(),
            "driver": driver["staff_no"] if driver else "",
            "accident_type": random.choice(ACCIDENT_TYPES),
            "severity": severity,
            "injuries": random.random() > 0.7,
            "police_notified": random.random() > 0.3,
            "third_party": random.random() > 0.5,
            "weather": random.choice(WEATHER_CONDITIONS),
            "road": random.choice(ROAD_CONDITIONS),
            "description": f"Accident involving {random.choice(['another vehicle', 'pedestrian', 'stationary object', 'animal'])}",
            "status": random.choice(["Reported", "Under Investigation", "Closed"])
        })
    return accidents

def generate_requisitions(staff, count=20):
    """Generate vehicle requisitions"""
    requisitions = []
    requesters = [s for s in staff if s["role"] in ["Driver", "Transport Supervisor", "Departmental Supervisor"]]
    
    for i in range(count):
        requester = random.choice(requesters) if requesters else random.choice(staff)
        
        travel_date = random_date(60, -30)  # Past and future dates
        return_date = (datetime.strptime(travel_date, '%Y-%m-%d') + timedelta(days=random.randint(1, 5))).strftime('%Y-%m-%d')
        
        requisitions.append({
            "requested_by": requester["staff_no"],
            "departure": random.choice(ROUTE_ORIGINS),
            "destination": random.choice(ROUTE_DESTINATIONS),
            "purpose": random.choice(["Client meeting", "Delivery", "Site inspection", "Training", "Conference", "Emergency"]),
            "travel_date": travel_date,
            "travel_time": f"{random.randint(6, 18):02d}:00",
            "return_date": return_date,
            "return_time": f"{random.randint(10, 20):02d}:00",
            "passengers": random.randint(1, 5),
            "passenger_names": "",
            "status": random.choice(["pending", "approved", "allocated", "completed", "rejected"])
        })
    return requisitions

# === EXCEL CREATION ===

def create_excel():
    print("🚀 Generating COMPLETE test data with Analytics-ready drivers...")
    
    staff = generate_staff(30)
    vehicles = generate_vehicles(25)
    routes = generate_routes(staff, vehicles, 60)
    fuel = generate_fuel(vehicles, 100)
    repairs = generate_repairs(vehicles, staff, 40)
    accidents = generate_accidents(staff, vehicles, 15)
    requisitions = generate_requisitions(staff, 20)
    
    wb = Workbook()
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # === 1. Staff Sheet ===
    ws_staff = wb.active
    ws_staff.title = "Staff"
    staff_headers = ["staff_no", "staff_name", "email", "phone", "designation", "department", "branch", "role"]
    ws_staff.append(staff_headers)
    for cell in ws_staff[1]:
        cell.fill = header_fill
        cell.font = header_font
    for s in staff:
        ws_staff.append([s[h] for h in staff_headers])
    
    # === 2. Fleet Sheet ===
    ws_veh = wb.create_sheet("Fleet")
    veh_headers = ["registration_num", "year_of_manufacture", "year_of_purchase", 
                   "make_model", "ownership", "department", "branch",
                   "minor_service_interval", "medium_service_interval", "major_service_interval",
                   "target_consumption_rate", "status", "current_mileage"]
    ws_veh.append(veh_headers)
    for cell in ws_veh[1]:
        cell.fill = header_fill
        cell.font = header_font
    for v in vehicles:
        ws_veh.append([v[h] for h in veh_headers])
    
    # === 3. Routes Sheet ===
    ws_routes = wb.create_sheet("Routes")
    route_headers = ["route_date", "route_name", "driver1", "vehicle",
                     "target_km", "actual_km", "target_fuel", "actual_fuel", "variance"]
    ws_routes.append(route_headers)
    for cell in ws_routes[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in routes:
        ws_routes.append([r[h] for h in route_headers])
    
    # === 4. Fuel Sheet ===
    ws_fuel = wb.create_sheet("TOTAL FUEL TEMPLATE")
    fuel_headers = ["department", "date", "registration", "card_num", "card_name",
                    "past", "current", "quantity", "amount", "place"]
    ws_fuel.append(fuel_headers)
    for cell in ws_fuel[1]:
        cell.fill = header_fill
        cell.font = header_font
    for f in fuel:
        ws_fuel.append([f[h] for h in fuel_headers])
    
    # === 5. Repairs Sheet ===
    ws_rep = wb.create_sheet("Repairs Template")
    rep_headers = ["date_in", "registration", "maintenance", "description",
                   "odometer", "technician", "garage"]
    ws_rep.append(rep_headers)
    for cell in ws_rep[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in repairs:
        ws_rep.append([r[h] for h in rep_headers])
    
    # === 6. Accidents Sheet ===
    ws_acc = wb.create_sheet("Accidents")
    acc_headers = ["case_number", "accident_date", "gps_location", "registration", 
                   "driver", "accident_type", "severity", "injuries", "police_notified",
                   "third_party", "weather", "road", "description", "status"]
    ws_acc.append(acc_headers)
    for cell in ws_acc[1]:
        cell.fill = header_fill
        cell.font = header_font
    for a in accidents:
        ws_acc.append([a[h] for h in acc_headers])
    
    # === 7. Requisitions Sheet ===
    ws_req = wb.create_sheet("Requisitions")
    req_headers = ["requested_by", "departure", "destination", "purpose",
                   "travel_date", "travel_time", "return_date", "return_time",
                   "passengers", "passenger_names", "status"]
    ws_req.append(req_headers)
    for cell in ws_req[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in requisitions:
        ws_req.append([r[h] for h in req_headers])
    
    # Auto-adjust columns
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(OUTPUT_FILE)
    
    # Print summary
    print(f"\n{'='*60}")
    print(f"✅ COMPLETE TEST DATA SAVED")
    print(f"{'='*60}")
    print(f"📁 File: {OUTPUT_FILE}")
    print(f"\n📊 RECORD COUNTS:")
    print(f"   • Staff: {len(staff)} (includes {len([s for s in staff if s['role']=='Driver'])} drivers)")
    print(f"   • Fleet: {len(vehicles)} vehicles")
    print(f"   • Routes: {len(routes)} operations")
    print(f"   • Fuel: {len(fuel)} records")
    print(f"   • Repairs: {len(repairs)} maintenance items")
    print(f"   • Accidents: {len(accidents)} incidents")
    print(f"   • Requisitions: {len(requisitions)} requests")
    print(f"\n📋 SHEETS: {', '.join(wb.sheetnames)}")
    print(f"\n✨ ANALYTICS READY:")
    print(f"   • Drivers linked to routes for productivity analytics")
    print(f"   • Drivers linked to accidents for safety scoring")
    print(f"   • Vehicles linked to fuel for consumption tracking")
    print(f"   • Vehicles linked to repairs for maintenance analytics")
    print(f"{'='*60}")
    
    return OUTPUT_FILE

if __name__ == "__main__":
    create_excel()
