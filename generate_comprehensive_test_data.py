#!/usr/bin/env python3
"""
Fleet Management System - Comprehensive Test Data Generator
Matches EXACT column names for Excel import
"""

import random
import uuid
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_FILE = "/root/.openclaw/workspace/fleet-system/fleet_test_data_comprehensive.xlsx"

# === DATA POOLS ===
FIRST_NAMES = ["John", "Mary", "James", "Patricia", "Robert", "Jennifer", "Michael", "Linda", 
               "William", "Elizabeth", "David", "Susan", "Richard", "Jessica", "Joseph", 
               "Sarah", "Thomas", "Karen", "Charles", "Nancy", "Daniel", "Lisa", "Matthew",
               "Betty", "Anthony", "Margaret", "Mark", "Sandra", "Donald", "Ashley"]

LAST_NAMES = ["Kamau", "Odhiambo", "Ochieng", "Wanjiku", "Kipchoge", "Mutua", "Omondi", 
              "Wanjiru", "Mwangi", "Kiptoo", "Korir", "Ruto", "Kiprotich", "Tanui",
              "Kiplagat", "Chebet", "Langat", "Kosgei", "Biwott", "Too"]

DEPARTMENTS = ["Transport", "Security", "Operations", "Finance", "HR", "IT", "Sales", "Logistics"]
BRANCHES = ["Nairobi HQ", "Mombasa", "Kisumu", "Nakuru", "Eldoret", "Thika", "Malindi"]
ROLES = ["Driver", "Transport Supervisor", "Departmental Supervisor", "Head of Department", "Security Personnel"]

VEHICLE_MAKES = ["Toyota Hilux", "Toyota Land Cruiser 79", "Toyota Land Cruiser 200", "Toyota Hiace",
                 "Toyota Coaster", "Toyota Corolla", "Nissan Patrol", "Nissan NP300", 
                 "Mitsubishi L200", "Ford Ranger", "Isuzu D-Max", "Mazda BT-50"]
OWNERSHIP_TYPES = ["Company", "Leased", "Hired"]
VEHICLE_STATUS = ["Active", "Under Maintenance", "Retired"]

ROUTE_ORIGINS = ["Nairobi HQ", "JKIA", "Industrial Area", "Westlands", "Mombasa Road Depot",
                 "Mombasa Port", "Mombasa CBD", "Changamwe", "Kisumu Depot", "Kisumu Port"]
ROUTE_DESTINATIONS = ["Mombasa", "Nairobi", "Kisumu", "Nakuru", "Eldoret", "Malindi", 
                      "Thika", "Namanga Border", "Busia Border", "Taveta"]

REPAIR_TYPES = ["Oil Change", "Brake Replacement", "Tire Rotation", "Engine Tune-up", 
                "Transmission Service", "Clutch Replacement", "Suspension Repair", 
                "Electrical Repair", "Air Conditioning", "Body Work"]
GARAGES = ["Toyota Kenya (Nairobi)", "Toyota Kenya (Mombasa)", "CMC Motors", "DT Dobie", 
           "City Garage", "Highway Service Center", "Kobil Workshop", "Shell Auto Care"]
FUEL_STATIONS = ["Shell", "Total", "Oilibya", "National Oil", "Kobil", "Rubis", "Delta"]

ACCIDENT_TYPES = ["Collision", "Pedestrian", "Property Damage", "Rollover", "Near Miss"]
ACCIDENT_SEVERITY = ["Minor", "Major", "Fatal"]
WEATHER_CONDITIONS = ["Clear", "Rainy", "Foggy", "Overcast"]
ROAD_CONDITIONS = ["Dry", "Wet", "Potholed", "Under Construction", "Muddy"]

def random_date(start_days=365, end_days=0):
    """Generate a random date between start_days ago and end_days ago"""
    days = random.randint(end_days, start_days)
    return (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

def random_datetime(start_days=365, end_days=0):
    """Generate a random datetime"""
    days = random.randint(end_days, start_days)
    hours = random.randint(6, 22)
    minutes = random.randint(0, 59)
    dt = datetime.now() - timedelta(days=days)
    dt = dt.replace(hour=hours, minute=minutes, second=0)
    return dt.strftime('%Y-%m-%d %H:%M')

def random_time():
    return f"{random.randint(6, 18):02d}:{random.randint(0, 59):02d}"

def generate_registration():
    """Generate Kenyan vehicle registration number"""
    prefixes = ['K', 'K', 'K', 'K', 'K', 'KB', 'KC', 'KX', 'KY']
    prefix = random.choice(prefixes)
    if len(prefix) == 1:
        letter = random.choice(['A', 'B', 'C', 'D', 'X', 'Y', 'Z'])
        numbers = random.randint(10, 999)
        suffix = random.choice(['A', 'B', 'C', 'X', 'Y'])
        suffix_num = random.randint(1000, 9999)
        return f"{prefix}{letter}{numbers} {suffix}{suffix_num}"
    else:
        numbers = random.randint(100, 999)
        suffix = random.choice(['A', 'B', 'C', 'X', 'Y'])
        suffix_num = random.randint(1000, 9999)
        return f"{prefix}{numbers} {suffix}{suffix_num}"

def generate_phone():
    """Generate Kenyan mobile number"""
    prefixes = ['07', '01']
    return f"{random.choice(prefixes)}{random.randint(10000000, 99999999)}"

# === GENERATOR FUNCTIONS ===

def generate_staff(count=25):
    """Generate staff records"""
    staff = []
    for i in range(count):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        dept = random.choice(DEPARTMENTS)
        role = random.choice(ROLES)
        
        staff.append({
            "staff_no": f"ST{1001+i}",
            "staff_name": f"{first} {last}",
            "email": f"{first.lower()}.{last.lower()}@company.com",
            "phone": generate_phone(),
            "designation": "Senior Driver" if role == "Driver" and random.random() > 0.5 else role,
            "department": dept,
            "branch": random.choice(BRANCHES),
            "role": role,
            "comments": ""
        })
    return staff

def generate_vehicles(count=20):
    """Generate vehicle fleet records"""
    vehicles = []
    for i in range(count):
        ownership = random.choice(OWNERSHIP_TYPES)
        status = random.choice(VEHICLE_STATUS)
        year_mfg = random.randint(2015, 2023)
        year_purchase = year_mfg + random.randint(0, 1)
        current_mileage = random.randint(15000, 180000)
        
        vehicles.append({
            "registration_num": generate_registration(),
            "year_of_manufacture": year_mfg,
            "year_of_purchase": year_purchase,
            "replacement_mileage": random.randint(200000, 350000),
            "replacement_age": random.randint(8, 12),
            "make_model": random.choice(VEHICLE_MAKES),
            "ownership": ownership,
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "minor_service_interval": 5000,
            "medium_service_interval": 15000,
            "major_service_interval": 30000,
            "target_consumption_rate": round(random.uniform(8.0, 14.0), 2),
            "status": status,
            "current_mileage": current_mileage,
            "last_service_date": random_date(120, 7),
            "next_service_due": random_date(90, -30)
        })
    return vehicles

def generate_routes(staff, vehicles, count=50):
    """Generate route/operations records"""
    routes = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    
    for i in range(count):
        driver = random.choice(drivers)
        vehicle = random.choice(vehicles)
        
        target_km = random.randint(100, 600)
        actual_km = target_km + random.randint(-30, 80)
        actual_km = max(50, actual_km)
        
        target_fuel = round(target_km / random.uniform(8.0, 12.0), 2)
        actual_fuel = round(actual_km / random.uniform(7.5, 13.0), 2)
        
        target_rate = round(random.uniform(8.0, 12.0), 2)
        actual_rate = round(actual_km / actual_fuel, 2) if actual_fuel > 0 else 0
        variance = round(actual_fuel - target_fuel, 2)
        
        origin = random.choice(ROUTE_ORIGINS)
        dest = random.choice(ROUTE_DESTINATIONS)
        while dest == origin:
            dest = random.choice(ROUTE_DESTINATIONS)
        
        routes.append({
            "route_date": random_date(90, 0),
            "route_name": f"{origin} - {dest}",
            "driver1_id": driver["staff_no"],
            "driver2_id": "",
            "co_driver_id": "",
            "vehicle_id": vehicle["registration_num"],
            "target_km": target_km,
            "actual_km": actual_km,
            "target_fuel_consumption": target_fuel,
            "actual_fuel": actual_fuel,
            "target_consumption_rate": target_rate,
            "actual_consumption_rate": actual_rate,
            "variance": variance,
            "comments": ""
        })
    return routes

def generate_fuel(vehicles, count=80):
    """Generate fuel records"""
    fuel = []
    for i in range(count):
        vehicle = random.choice(vehicles)
        current = vehicle["current_mileage"] - random.randint(1000, 15000) + random.randint(0, 5000)
        past = current - random.randint(300, 2500)
        
        distance = current - past
        km_per_l = random.uniform(7.5, 13.0)
        qty = round(distance / km_per_l, 2)
        price_per_l = random.uniform(175, 198)
        amt = round(qty * price_per_l, 2)
        
        fuel.append({
            "department": vehicle["department"],
            "fuel_date": random_date(120, 0),
            "vehicle_id": vehicle["registration_num"],
            "card_num": f"FC{random.randint(100000, 999999)}",
            "card_name": random.choice(FUEL_STATIONS),
            "past_mileage": past,
            "current_mileage": current,
            "quantity_liters": qty,
            "amount": amt,
            "place": random.choice(BRANCHES)
        })
    return fuel

def generate_repairs(vehicles, staff, count=35):
    """Generate repair/maintenance records"""
    repairs = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    
    for i in range(count):
        vehicle = random.choice(vehicles)
        driver = random.choice(drivers)
        date_in = random_date(180, 0)
        
        target_hours = round(random.uniform(2, 24), 1)
        actual_hours = round(target_hours + random.uniform(-1, 8), 1)
        actual_hours = max(1, actual_hours)
        productivity = round(target_hours / actual_hours, 2)
        
        is_pm = random.random() > 0.3
        pm_type = random.choice(REPAIR_TYPES) if is_pm else ""
        breakdown = "" if is_pm else random.choice(["Engine overheating", "Brake failure", 
                                                     "Electrical fault", "Suspension noise"])
        
        repairs.append({
            "date_in": date_in,
            "vehicle_id": vehicle["registration_num"],
            "preventative_maintenance": pm_type,
            "breakdown_description": breakdown,
            "odometer_reading": vehicle["current_mileage"],
            "driver_id": driver["staff_name"],
            "assigned_technician": random.choice(["John Mechanic", "Peter Fixer", "James Repair", 
                                                   "Ali Technician", "Hassan Auto"]),
            "date_out": (datetime.strptime(date_in, '%Y-%m-%d') + timedelta(days=random.randint(1,7))).strftime('%Y-%m-%d'),
            "actual_repair_hours": actual_hours,
            "target_repair_hours": target_hours,
            "productivity_ratio": productivity,
            "garage_name": random.choice(GARAGES),
            "cost": round(random.uniform(3000, 75000), 2),
            "status": random.choice(["Completed", "In Progress", "Pending"])
        })
    return repairs

def generate_accidents(staff, vehicles, count=10):
    """Generate accident investigation records"""
    accidents = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    
    for i in range(count):
        driver = random.choice(drivers)
        vehicle = random.choice(vehicles)
        severity = random.choice(ACCIDENT_SEVERITY)
        
        accidents.append({
            "case_number": f"ACC-2025-{1001+i:04d}",
            "accident_date": random_datetime(180, 0),
            "gps_location": f"{random.uniform(-1.5, -1.0):.4f}, {random.uniform(36.7, 37.1):.4f}",
            "vehicle_id": vehicle["registration_num"],
            "driver_id": driver["staff_no"],
            "accident_type": random.choice(ACCIDENT_TYPES),
            "severity": severity,
            "injuries_reported": random.choice([True, False]),
            "police_notified": True if severity in ["Major", "Fatal"] else random.choice([True, False]),
            "third_party_involved": random.choice([True, False]),
            "weather_condition": random.choice(WEATHER_CONDITIONS),
            "road_condition": random.choice(ROAD_CONDITIONS),
            "incident_description": f"Accident occurred on {random.choice(['highway', 'urban road', 'rural road'])}. " +
                                    f"Vehicle {random.choice(['collided with', 'skidded on', 'hit'])} " +
                                    random.choice(["another vehicle", "road barrier", "pothole", "pedestrian"]),
            "status": random.choice(["Reported", "Under Investigation", "Root Cause Identified", "CAPA In Progress", "Closed"]),
            "reported_by": random.choice(staff)["staff_no"]
        })
    return accidents

def generate_requisitions(staff, count=15):
    """Generate vehicle requisition records"""
    requisitions = []
    requesters = [s for s in staff if s["role"] != "Driver"]
    
    for i in range(count):
        requester = random.choice(requesters)
        travel_date = random_date(60, -30)  # Some past, some future
        
        origin = random.choice(ROUTE_ORIGINS[:5])
        dest = random.choice(ROUTE_DESTINATIONS[:6])
        
        requisitions.append({
            "requested_by": requester["staff_no"],
            "place_of_departure": origin,
            "destination": dest,
            "purpose": random.choice(["Client meeting", "Site inspection", "Training", 
                                      "Conference", "Delivery", "Emergency response"]),
            "travel_date": travel_date,
            "travel_time": random_time(),
            "return_date": (datetime.strptime(travel_date, '%Y-%m-%d') + timedelta(days=random.randint(0, 3))).strftime('%Y-%m-%d'),
            "return_time": random_time(),
            "num_passengers": random.randint(1, 5),
            "passenger_names": requester["staff_name"],
            "status": random.choice(["Draft", "Submitted", "Manager Approved", "Transport Approved", 
                                    "Allocated", "Completed", "Rejected"])
        })
    return requisitions

# === EXCEL CREATION ===

def create_excel():
    print("🚀 Generating comprehensive test data...")
    
    staff = generate_staff(25)
    vehicles = generate_vehicles(20)
    routes = generate_routes(staff, vehicles, 50)
    fuel = generate_fuel(vehicles, 80)
    repairs = generate_repairs(vehicles, staff, 35)
    accidents = generate_accidents(staff, vehicles, 10)
    requisitions = generate_requisitions(staff, 15)
    
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # === 1. STAFF Sheet ===
    ws_staff = wb.active
    ws_staff.title = "Staff"
    staff_headers = ["staff_no", "staff_name", "email", "phone", "designation", 
                     "department", "branch", "role", "comments"]
    ws_staff.append(staff_headers)
    for cell in ws_staff[1]:
        cell.fill = header_fill
        cell.font = header_font
    for s in staff:
        ws_staff.append([s[h] for h in staff_headers])
    
    # === 2. VEHICLES Sheet (Renamed to Fleet for import) ===
    ws_veh = wb.create_sheet("Fleet")
    veh_headers = ["registration_num", "year_of_manufacture", "year_of_purchase", 
                   "replacement_mileage", "replacement_age", "make_model", "ownership", 
                   "department", "branch", "minor_service_interval", "medium_service_interval", 
                   "major_service_interval", "target_consumption_rate", "status", 
                   "current_mileage", "last_service_date", "next_service_due"]
    ws_veh.append(veh_headers)
    for cell in ws_veh[1]:
        cell.fill = header_fill
        cell.font = header_font
    for v in vehicles:
        ws_veh.append([v[h] for h in veh_headers])
    
    # === 3. ROUTES Sheet ===
    ws_routes = wb.create_sheet("Routes")
    route_headers = ["route_date", "route_name", "driver1_id", "driver2_id", "co_driver_id", 
                     "vehicle_id", "target_km", "actual_km", "target_fuel_consumption", 
                     "actual_fuel", "target_consumption_rate", "actual_consumption_rate", 
                     "variance", "comments"]
    ws_routes.append(route_headers)
    for cell in ws_routes[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in routes:
        ws_routes.append([r[h] for h in route_headers])
    
    # === 4. FUEL Sheet (Named as expected by system) ===
    ws_fuel = wb.create_sheet("TOTAL FUEL TEMPLATE")
    fuel_headers = ["department", "fuel_date", "vehicle_id", "card_num", "card_name",
                    "past_mileage", "current_mileage", "quantity_liters", "amount", "place"]
    ws_fuel.append(fuel_headers)
    for cell in ws_fuel[1]:
        cell.fill = header_fill
        cell.font = header_font
    for f in fuel:
        ws_fuel.append([f[h] for h in fuel_headers])
    
    # === 5. REPAIRS Sheet (Named as expected by system) ===
    ws_rep = wb.create_sheet("Repairs Template")
    rep_headers = ["date_in", "vehicle_id", "preventative_maintenance", "breakdown_description",
                   "odometer_reading", "driver_id", "assigned_technician", "date_out",
                   "actual_repair_hours", "target_repair_hours", "productivity_ratio",
                   "garage_name", "cost", "status"]
    ws_rep.append(rep_headers)
    for cell in ws_rep[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in repairs:
        ws_rep.append([r[h] for h in rep_headers])
    
    # === 6. ACCIDENTS Sheet ===
    ws_acc = wb.create_sheet("Accidents")
    acc_headers = ["case_number", "accident_date", "gps_location", "vehicle_id", "driver_id",
                   "accident_type", "severity", "injuries_reported", "police_notified",
                   "third_party_involved", "weather_condition", "road_condition",
                   "incident_description", "status", "reported_by"]
    ws_acc.append(acc_headers)
    for cell in ws_acc[1]:
        cell.fill = header_fill
        cell.font = header_font
    for a in accidents:
        ws_acc.append([a[h] for h in acc_headers])
    
    # === 7. REQUISITIONS Sheet ===
    ws_req = wb.create_sheet("Requisitions")
    req_headers = ["requested_by", "place_of_departure", "destination", "purpose", 
                   "travel_date", "travel_time", "return_date", "return_time",
                   "num_passengers", "passenger_names", "status"]
    ws_req.append(req_headers)
    for cell in ws_req[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in requisitions:
        ws_req.append([r[h] for h in req_headers])
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Test data saved to: {OUTPUT_FILE}")
    print(f"\n📊 Summary:")
    print(f"   • Staff: {len(staff)} records")
    print(f"   • Fleet: {len(vehicles)} vehicles")
    print(f"   • Routes: {len(routes)} operations")
    print(f"   • Fuel: {len(fuel)} records")
    print(f"   • Repairs: {len(repairs)} maintenance items")
    print(f"   • Accidents: {len(accidents)} cases")
    print(f"   • Requisitions: {len(requisitions)} requests")
    print(f"\n📁 Sheets: {', '.join(wb.sheetnames)}")
    return OUTPUT_FILE

if __name__ == "__main__":
    create_excel()
