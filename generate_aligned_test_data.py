#!/usr/bin/env python3
"""
Fleet Management System - ALIGNED Test Data Generator
Column names EXACTLY match what backend/upload.ts expects
"""

import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUTPUT_FILE = "/root/.openclaw/workspace/fleet-system/fleet_test_data_aligned.xlsx"

# === DATA POOLS ===
FIRST_NAMES = ["John", "Mary", "James", "Patricia", "Robert", "Jennifer", "Michael", "Linda", 
               "William", "Elizabeth", "David", "Susan", "Richard", "Jessica", "Joseph", 
               "Sarah", "Thomas", "Karen", "Charles", "Nancy", "Daniel", "Lisa", "Matthew",
               "Betty", "Anthony", "Margaret", "Mark", "Sandra", "Donald", "Ashley"]

LAST_NAMES = ["Kamau", "Odhiambo", "Ochieng", "Wanjiku", "Kipchoge", "Mutua", "Omondi", 
              "Wanjiru", "Mwangi", "Kiptoo", "Korir", "Ruto", "Kiprotich", "Tanui",
              "Kiplagat", "Chebet", "Langat", "Kosgei", "Biwott", "Too"]

DEPARTMENTS = ["Transport", "Security", "Operations", "Finance", "HR", "IT", "Sales", "Logistics"]
BRANCHES = ["Nairobi HQ", "Mombasa", "Kisumu", "Nakuru", "Eldoret", "Thika", "Malindi"]
ROLES = ["Driver", "Transport Supervisor", "Departmental Supervisor", "Head of Department", "Security Personnel"]

VEHICLE_MAKES = ["Toyota Hilux", "Toyota Land Cruiser 79", "Toyota Land Cruiser 200", "Toyota Hiace",
                 "Toyota Coaster", "Toyota Corolla", "Nissan Patrol", "Nissan NP300", 
                 "Mitsubishi L200", "Ford Ranger", "Isuzu D-Max", "Mazda BT-50"]
OWNERSHIP_TYPES = ["Company", "Leased", "Hired"]
VEHICLE_STATUS = ["Active", "Under Maintenance", "Retired"]

REPAIR_TYPES = ["Oil Change", "Brake Replacement", "Tire Rotation", "Engine Tune-up", 
                "Transmission Service", "Clutch Replacement", "Suspension Repair", 
                "Electrical Repair", "Air Conditioning", "Body Work"]
GARAGES = ["Toyota Kenya (Nairobi)", "Toyota Kenya (Mombasa)", "CMC Motors", "DT Dobie", 
           "City Garage", "Highway Service Center", "Kobil Workshop", "Shell Auto Care"]
FUEL_STATIONS = ["Shell", "Total", "Oilibya", "National Oil", "Kobil", "Rubis", "Delta"]

ROUTE_ORIGINS = ["Nairobi HQ", "JKIA", "Industrial Area", "Westlands", "Mombasa Road Depot",
                 "Mombasa Port", "Mombasa CBD", "Changamwe", "Kisumu Depot", "Kisumu Port"]
ROUTE_DESTINATIONS = ["Mombasa", "Nairobi", "Kisumu", "Nakuru", "Eldoret", "Malindi", 
                      "Thika", "Namanga Border", "Busia Border", "Taveta"]

def random_date(start_days=365, end_days=0):
    days = random.randint(end_days, start_days)
    return (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

def generate_registration():
    prefixes = ['K', 'K', 'K', 'K', 'K', 'KB', 'KC', 'KX', 'KY']
    prefix = random.choice(prefixes)
    if len(prefix) == 1:
        letter = random.choice(['A', 'B', 'C', 'D', 'X', 'Y', 'Z'])
        numbers = random.randint(10, 999)
        suffix = random.choice(['A', 'B', 'C', 'X', 'Y'])
        suffix_num = random.randint(1000, 9999)
        return f"{prefix}{letter}{numbers} {suffix}{suffix_num}"
    else:
        numbers = random.randint(100, 999)
        suffix = random.choice(['A', 'B', 'C', 'X', 'Y'])
        suffix_num = random.randint(1000, 9999)
        return f"{prefix}{numbers} {suffix}{suffix_num}"

def generate_phone():
    prefixes = ['07', '01']
    return f"{random.choice(prefixes)}{random.randint(10000000, 99999999)}"

# === GENERATOR FUNCTIONS ===

def generate_staff(count=25):
    staff = []
    for i in range(count):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        staff.append({
            "staff_no": f"ST{1001+i}",
            "name": f"{first} {last}",  # Changed from 'staff_name' to 'name' for backend
            "email": f"{first.lower()}.{last.lower()}@company.com",
            "phone": generate_phone(),
            "designation": random.choice(["Senior Driver", "Driver", "Supervisor", "Manager"]),
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "role": random.choice(ROLES)
        })
    return staff

def generate_vehicles(count=20):
    vehicles = []
    for i in range(count):
        ownership = random.choice(OWNERSHIP_TYPES)
        year_mfg = random.randint(2015, 2023)
        current_mileage = random.randint(15000, 180000)
        
        vehicles.append({
            "registration": generate_registration(),  # Backend expects 'registration'
            "year_of_manufacture": year_mfg,
            "year_of_purchase": year_mfg + random.randint(0, 1),
            "make_model": random.choice(VEHICLE_MAKES),
            "ownership": ownership,
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "minor_service_interval": 5000,
            "medium_service_interval": 15000,
            "major_service_interval": 30000,
            "consumption_rate": round(random.uniform(8.0, 14.0), 2),  # Backend looks for 'consumption'
            "status": random.choice(VEHICLE_STATUS),
            "mileage": current_mileage  # Backend expects 'mileage'
        })
    return vehicles

def generate_routes(staff, vehicles, count=50):
    routes = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    
    for i in range(count):
        driver = random.choice(drivers)
        vehicle = random.choice(vehicles)
        
        target_km = random.randint(100, 600)
        actual_km = max(50, target_km + random.randint(-30, 80))
        target_fuel = round(target_km / random.uniform(8.0, 12.0), 2)
        actual_fuel = round(actual_km / random.uniform(7.5, 13.0), 2)
        variance = round(actual_fuel - target_fuel, 2)
        
        origin = random.choice(ROUTE_ORIGINS)
        dest = random.choice(ROUTE_DESTINATIONS)
        while dest == origin:
            dest = random.choice(ROUTE_DESTINATIONS)
        
        routes.append({
            "route_date": random_date(90, 0),
            "route_name": f"{origin} - {dest}",
            "driver1": driver["staff_no"],  # Backend expects 'driver1'
            "vehicle": vehicle["registration"],  # Backend expects 'vehicle'
            "target_km": target_km,
            "actual_km": actual_km,
            "target_fuel": target_fuel,
            "actual_fuel": actual_fuel,
            "variance": variance
        })
    return routes

def generate_fuel(vehicles, count=80):
    fuel = []
    for i in range(count):
        vehicle = random.choice(vehicles)
        current = random.randint(20000, 180000)
        past = current - random.randint(300, 2500)
        
        distance = current - past
        km_per_l = random.uniform(7.5, 13.0)
        qty = round(distance / km_per_l, 2)
        price_per_l = random.uniform(175, 198)
        amt = round(qty * price_per_l, 2)
        
        fuel.append({
            "department": vehicle["department"],
            "date": random_date(120, 0),  # Backend expects 'date'
            "registration": vehicle["registration"],  # Backend expects 'registration' NOT 'vehicle_id'
            "card_num": f"FC{random.randint(100000, 999999)}",
            "card_name": random.choice(FUEL_STATIONS),
            "past": past,  # Backend expects 'past'
            "current": current,  # Backend expects 'current'
            "quantity": qty,  # Backend expects 'quantity'
            "amount": amt,
            "place": random.choice(BRANCHES)
        })
    return fuel

def generate_repairs(vehicles, staff, count=35):
    repairs = []
    drivers = [s for s in staff if s["role"] == "Driver"]
    
    for i in range(count):
        vehicle = random.choice(vehicles)
        driver = random.choice(drivers) if drivers else None
        
        repairs.append({
            "date_in": random_date(180, 0),
            "registration": vehicle["registration"],  # Backend expects 'registration' NOT 'vehicle_id'
            "maintenance": random.choice(REPAIR_TYPES),  # Backend expects 'maintenance'
            "description": random.choice(["Routine service", "Minor repair", "Annual check"]),
            "odometer": random.randint(20000, 180000),  # Backend expects 'odometer'
            "technician": random.choice(["John Mechanic", "Peter Fixer", "James Repair"]),  # Backend expects 'technician'
            "garage": random.choice(GARAGES)  # Backend expects 'garage'
        })
    return repairs

# === EXCEL CREATION ===

def create_excel():
    print("🚀 Generating ALIGNED test data...")
    
    staff = generate_staff(25)
    vehicles = generate_vehicles(20)
    routes = generate_routes(staff, vehicles, 50)
    fuel = generate_fuel(vehicles, 80)
    repairs = generate_repairs(vehicles, staff, 35)
    
    wb = Workbook()
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # === 1. Staff Sheet ===
    ws_staff = wb.active
    ws_staff.title = "Staff"
    staff_headers = ["staff_no", "name", "email", "phone", "designation", "department", "branch", "role"]
    ws_staff.append(staff_headers)
    for cell in ws_staff[1]:
        cell.fill = header_fill
        cell.font = header_font
    for s in staff:
        ws_staff.append([s[h] for h in staff_headers])
    
    # === 2. Fleet Sheet (Backend looks for 'vehicles' or 'fleet') ===
    ws_veh = wb.create_sheet("Fleet")
    veh_headers = ["registration", "year_of_manufacture", "year_of_purchase", 
                   "make_model", "ownership", "department", "branch",
                   "minor_service_interval", "medium_service_interval", "major_service_interval",
                   "consumption_rate", "status", "mileage"]
    ws_veh.append(veh_headers)
    for cell in ws_veh[1]:
        cell.fill = header_fill
        cell.font = header_font
    for v in vehicles:
        ws_veh.append([v[h] for h in veh_headers])
    
    # === 3. Routes Sheet ===
    ws_routes = wb.create_sheet("Routes")
    route_headers = ["route_date", "route_name", "driver1", "vehicle",
                     "target_km", "actual_km", "target_fuel", "actual_fuel", "variance"]
    ws_routes.append(route_headers)
    for cell in ws_routes[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in routes:
        ws_routes.append([r[h] for h in route_headers])
    
    # === 4. Fuel Sheet (Backend looks for 'fuel' or 'total fuel template') ===
    ws_fuel = wb.create_sheet("TOTAL FUEL TEMPLATE")
    fuel_headers = ["department", "date", "registration", "card_num", "card_name",
                    "past", "current", "quantity", "amount", "place"]
    ws_fuel.append(fuel_headers)
    for cell in ws_fuel[1]:
        cell.fill = header_fill
        cell.font = header_font
    for f in fuel:
        ws_fuel.append([f[h] for h in fuel_headers])
    
    # === 5. Repairs Sheet (Backend looks for 'repairs' or 'repairs template') ===
    ws_rep = wb.create_sheet("Repairs Template")
    rep_headers = ["date_in", "registration", "maintenance", "description",
                   "odometer", "technician", "garage"]
    ws_rep.append(rep_headers)
    for cell in ws_rep[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in repairs:
        ws_rep.append([r[h] for h in rep_headers])
    
    # Auto-adjust columns
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Aligned test data saved to: {OUTPUT_FILE}")
    print(f"\n📊 Summary:")
    print(f"   • Staff: {len(staff)} records")
    print(f"   • Fleet: {len(vehicles)} vehicles")
    print(f"   • Routes: {len(routes)} operations")
    print(f"   • Fuel: {len(fuel)} records")
    print(f"   • Repairs: {len(repairs)} maintenance items")
    print(f"\n📁 Sheets: {', '.join(wb.sheetnames)}")
    print(f"\n⚠️  NOTE: Fuel and Repairs now use 'registration' column")
    print(f"    (vehicle plate number) instead of 'vehicle_id' for proper lookup")
    return OUTPUT_FILE

if __name__ == "__main__":
    create_excel()
