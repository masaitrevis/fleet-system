#!/usr/bin/env python3
"""
Generate comprehensive test data for Fleet Management System
Aligned with PostgreSQL schema
"""

import random
import uuid
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Configuration
OUTPUT_FILE = "/tmp/fleet_test_data_complete.xlsx"

# Random data generators
def random_date(start_days_ago=365, end_days_ago=0):
    """Generate random date within range"""
    days_ago = random.randint(end_days_ago, start_days_ago)
    return (datetime.now() - timedelta(days=days_ago)).strftime('%Y-%m-%d')

def random_time():
    """Generate random time"""
    hour = random.randint(6, 18)
    minute = random.randint(0, 59)
    return f"{hour:02d}:{minute:02d}"

# Sample data pools
FIRST_NAMES = ["John", "Mary", "James", "Patricia", "Robert", "Jennifer", "Michael", "Linda", "William", "Elizabeth",
               "David", "Susan", "Richard", "Jessica", "Joseph", "Sarah", "Thomas", "Karen", "Charles", "Nancy"]

LAST_NAMES = ["Kamau", "Odhiambo", "Ochieng", "Wanjiku", "Kipchoge", "Mutua", "Omondi", "Wanjiru", "Mwangi", "Kiptoo",
              "Cheruiyot", "Langat", "Korir", "Rotich", "Kosgei", "Jepchirchir", "Kiplagat", "Chepkoech", "Jeptoo", "Biwott"]

DEPARTMENTS = ["Transport", "Security", "Operations", "Finance", "HR", "IT", "Sales", "Logistics", "Administration"]

BRANCHES = ["Nairobi HQ", "Mombasa", "Kisumu", "Nakuru", "Eldoret", "Thika", "Machakos", "Nyeri"]

ROLES = ["Driver", "Transport Supervisor", "Departmental Supervisor", "Head of Department", "Security Personnel", "Manager"]

VEHICLE_MAKES = ["Toyota Hilux", "Toyota Land Cruiser", "Toyota Hiace", "Nissan Patrol", "Nissan Navara",
                 "Mitsubishi L200", "Ford Ranger", "Isuzu D-Max", "Mazda BT-50", "Volkswagen Amarok"]

OWNERSHIP_TYPES = ["Company", "Leased", "Hired"]

VEHICLE_STATUSES = ["Active", "Under Maintenance", "Retired"]

ROUTE_NAMES = ["Nairobi-Mombasa Highway", "Nairobi-Nakuru Highway", "Nairobi-Thika Road", "CBD Round Trip",
               "Airport Run", "Warehouse Delivery", "Client Visit", "Field Survey", "Security Patrol", "Emergency Response"]

REPAIR_TYPES = ["Oil Change", "Brake Replacement", "Tire Rotation", "Engine Tune-up", "Transmission Service",
                "Air Conditioning Repair", "Electrical System Check", "Suspension Repair", "Battery Replacement", "General Service"]

GARAGES = ["Toyota Kenya", "CMC Motors", "DT Dobie", "General Motors", "City Garage", "Highway Service Center",
           "Auto Express", "Quick Fit Autocare", "Speedy Motors", "Elite Auto Services"]

FUEL_STATIONS = ["Shell", "Total", "Oilibya", "National Oil", "Kobil", "Rubis", "Hass", "Gulf Energy"]

PLACES = ["Nairobi", "Mombasa", "Kisumu", "Nakuru", "Eldoret", "Thika", "Machakos", "Nyeri", "Meru", "Nanyuki"]

def generate_staff_data(count=20):
    """Generate staff data"""
    staff = []
    for i in range(count):
        first_name = random.choice(FIRST_NAMES)
        last_name = random.choice(LAST_NAMES)
        staff.append({
            "id": str(uuid.uuid4()),
            "staff_no": f"ST{1000 + i:04d}",
            "staff_name": f"{first_name} {last_name}",
            "email": f"{first_name.lower()}.{last_name.lower()}@g4s.com",
            "phone": f"07{random.randint(10000000, 99999999)}",
            "designation": random.choice(["Senior Driver", "Driver", "Supervisor", "Assistant"]),
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "role": random.choice(ROLES),
            "comments": ""
        })
    return staff

def generate_vehicle_data(count=15):
    """Generate vehicle data"""
    vehicles = []
    for i in range(count):
        make_model = random.choice(VEHICLE_MAKES)
        year_manufacture = random.randint(2015, 2023)
        ownership = random.choice(OWNERSHIP_TYPES)
        
        # Service intervals based on ownership type
        if ownership == "Company":
            minor_interval = 5000
            medium_interval = 15000
            major_interval = 30000
        else:
            minor_interval = 3000
            medium_interval = 10000
            major_interval = 20000
            
        vehicles.append({
            "id": str(uuid.uuid4()),
            "registration_num": f"K{random.choice(['A', 'B', 'C', 'D', 'X', 'Y'])}{random.randint(100, 999)} {random.choice(['A', 'B', 'C', 'D', 'X', 'Y'])}{random.randint(1000, 9999)}",
            "year_of_manufacture": year_manufacture,
            "year_of_purchase": year_manufacture + random.randint(0, 1),
            "replacement_mileage": random.randint(150000, 300000),
            "replacement_age": random.randint(7, 12),
            "make_model": make_model,
            "ownership": ownership,
            "department": random.choice(DEPARTMENTS),
            "branch": random.choice(BRANCHES),
            "minor_service_interval": minor_interval,
            "medium_service_interval": medium_interval,
            "major_service_interval": major_interval,
            "target_consumption_rate": round(random.uniform(8.0, 15.0), 2),
            "status": random.choice(VEHICLE_STATUSES),
            "current_mileage": random.randint(10000, 150000),
            "last_service_date": random_date(180, 30),
            "next_service_due": random_date(30, -30)
        })
    return vehicles

def generate_fuel_data(vehicles, count=50):
    """Generate fuel records"""
    fuel_records = []
    for i in range(count):
        vehicle = random.choice(vehicles)
        past_mileage = vehicle["current_mileage"] - random.randint(500, 3000)
        current_mileage = past_mileage + random.randint(400, 2500)
        distance = current_mileage - past_mileage
        quantity = round(distance / random.uniform(8.0, 12.0), 2)
        amount = round(quantity * random.uniform(170, 195), 2)
        
        fuel_records.append({
            "department": vehicle["department"],
            "fuel_date": random_date(90, 0),
            "vehicle_id": vehicle["id"],
            "registration_num": vehicle["registration_num"],
            "card_num": f"FC{random.randint(100000, 999999)}",
            "card_name": random.choice(FUEL_STATIONS),
            "past_mileage": past_mileage,
            "current_mileage": current_mileage,
            "quantity_liters": quantity,
            "amount": amount,
            "place": random.choice(PLACES)
        })
    return fuel_records

def generate_route_data(staff, vehicles, count=30):
    """Generate route/operations data"""
    routes = []
    for i in range(count):
        driver = random.choice([s for s in staff if s["role"] == "Driver"])
        vehicle = random.choice(vehicles)
        target_km = random.randint(100, 800)
        actual_km = target_km + random.randint(-50, 100)
        target_fuel = round(target_km / random.uniform(8.0, 12.0), 2)
        actual_fuel = round(actual_km / random.uniform(7.5, 13.0), 2)
        
        routes.append({
            "route_date": random_date(60, 0),
            "route_name": random.choice(ROUTE_NAMES),
            "driver1_id": driver["id"],
            "driver1_name": driver["staff_name"],
            "vehicle_id": vehicle["id"],
            "registration_num": vehicle["registration_num"],
            "target_km": target_km,
            "actual_km": actual_km,
            "target_fuel_consumption": target_fuel,
            "actual_fuel": actual_fuel,
            "variance": round(actual_fuel - target_fuel, 2),
            "comments": ""
        })
    return routes

def generate_repair_data(vehicles, staff, count=25):
    """Generate repair/maintenance data"""
    repairs = []
    for i in range(count):
        vehicle = random.choice(vehicles)
        driver = random.choice([s for s in staff if s["role"] == "Driver"])
        date_in = random_date(120, 0)
        target_hours = random.uniform(2, 24)
        actual_hours = target_hours + random.uniform(-2, 8)
        
        repairs.append({
            "date_in": date_in,
            "vehicle_id": vehicle["id"],
            "registration_num": vehicle["registration_num"],
            "preventative_maintenance": random.choice(REPAIR_TYPES),
            "breakdown_description": "Routine maintenance" if random.random() > 0.3 else "Minor repair required",
            "odometer_reading": vehicle["current_mileage"],
            "driver_id": driver["id"],
            "driver_name": driver["staff_name"],
            "assigned_technician": random.choice(["John Mechanic", "Peter Fixer", "James Repair", "Michael Tech"]),
            "date_out": (datetime.strptime(date_in, '%Y-%m-%d') + timedelta(days=random.randint(1, 5))).strftime('%Y-%m-%d'),
            "actual_repair_hours": round(actual_hours, 2),
            "target_repair_hours": round(target_hours, 2),
            "productivity_ratio": round(target_hours / actual_hours, 2) if actual_hours > 0 else 1.0,
            "garage_name": random.choice(GARAGES),
            "cost": round(random.uniform(5000, 50000), 2),
            "status": random.choice(["Completed", "In Progress", "Pending"])
        })
    return repairs

def generate_requisition_data(staff, count=10):
    """Generate vehicle requisition data"""
    requisitions = []
    statuses = ["pending", "approved", "allocated", "completed"]
    
    for i in range(count):
        requester = random.choice(staff)
        from_place = random.choice(PLACES)
        to_place = random.choice([p for p in PLACES if p != from_place])
        travel_date = random_date(30, -7)
        
        requisitions.append({
            "request_no": f"REQ-2025{random.randint(1, 12):02d}-{random.randint(1000, 9999)}",
            "requested_by": requester["id"],
            "requester_name": requester["staff_name"],
            "place_of_departure": from_place,
            "destination": to_place,
            "purpose": random.choice(["Client meeting", "Site inspection", "Training", "Emergency response", "Equipment delivery"]),
            "travel_date": travel_date,
            "travel_time": random_time(),
            "return_date": (datetime.strptime(travel_date, '%Y-%m-%d') + timedelta(days=random.randint(0, 3))).strftime('%Y-%m-%d'),
            "return_time": random_time(),
            "num_passengers": random.randint(1, 5),
            "passenger_names": "",
            "status": random.choice(statuses)
        })
    return requisitions

def create_excel_file():
    """Create comprehensive Excel file with all test data"""
    print("🚀 Generating test data...")
    
    # Generate data
    staff = generate_staff_data(20)
    vehicles = generate_vehicle_data(15)
    fuel_records = generate_fuel_data(vehicles, 50)
    routes = generate_route_data(staff, vehicles, 30)
    repairs = generate_repair_data(vehicles, staff, 25)
    requisitions = generate_requisition_data(staff, 10)
    
    # Create workbook
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Staff Sheet
    ws_staff = wb.active
    ws_staff.title = "Staff"
    staff_headers = ["staff_no", "staff_name", "email", "phone", "designation", "department", "branch", "role", "comments"]
    ws_staff.append(staff_headers)
    for cell in ws_staff[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for s in staff:
        ws_staff.append([s["staff_no"], s["staff_name"], s["email"], s["phone"], 
                        s["designation"], s["department"], s["branch"], s["role"], s["comments"]])
    
    # 2. Vehicles Sheet
    ws_vehicles = wb.create_sheet("Vehicles")
    vehicle_headers = ["registration_num", "year_of_manufacture", "year_of_purchase", "replacement_mileage",
                      "replacement_age", "make_model", "ownership", "department", "branch",
                      "minor_service_interval", "medium_service_interval", "major_service_interval",
                      "target_consumption_rate", "status", "current_mileage", "last_service_date", "next_service_due"]
    ws_vehicles.append(vehicle_headers)
    for cell in ws_vehicles[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for v in vehicles:
        ws_vehicles.append([v["registration_num"], v["year_of_manufacture"], v["year_of_purchase"],
                           v["replacement_mileage"], v["replacement_age"], v["make_model"], v["ownership"],
                           v["department"], v["branch"], v["minor_service_interval"], v["medium_service_interval"],
                           v["major_service_interval"], v["target_consumption_rate"], v["status"],
                           v["current_mileage"], v["last_service_date"], v["next_service_due"]])
    
    # 3. Fuel Records Sheet
    ws_fuel = wb.create_sheet("Fuel")
    fuel_headers = ["department", "fuel_date", "registration_num", "card_num", "card_name",
                   "past_mileage", "current_mileage", "quantity_liters", "amount", "place"]
    ws_fuel.append(fuel_headers)
    for cell in ws_fuel[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for f in fuel_records:
        ws_fuel.append([f["department"], f["fuel_date"], f["registration_num"], f["card_num"],
                       f["card_name"], f["past_mileage"], f["current_mileage"], f["quantity_liters"],
                       f["amount"], f["place"]])
    
    # 4. Routes Sheet
    ws_routes = wb.create_sheet("Routes")
    route_headers = ["route_date", "route_name", "driver1_name", "registration_num", "target_km",
                    "actual_km", "target_fuel_consumption", "actual_fuel", "variance", "comments"]
    ws_routes.append(route_headers)
    for cell in ws_routes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for r in routes:
        ws_routes.append([r["route_date"], r["route_name"], r["driver1_name"], r["registration_num"],
                         r["target_km"], r["actual_km"], r["target_fuel_consumption"], r["actual_fuel"],
                         r["variance"], r["comments"]])
    
    # 5. Repairs Sheet
    ws_repairs = wb.create_sheet("Repairs")
    repair_headers = ["date_in", "registration_num", "preventative_maintenance", "breakdown_description",
                     "odometer_reading", "driver_name", "assigned_technician", "date_out",
                     "actual_repair_hours", "target_repair_hours", "productivity_ratio",
                     "garage_name", "cost", "status"]
    ws_repairs.append(repair_headers)
    for cell in ws_repairs[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for r in repairs:
        ws_repairs.append([r["date_in"], r["registration_num"], r["preventative_maintenance"],
                          r["breakdown_description"], r["odometer_reading"], r["driver_name"],
                          r["assigned_technician"], r["date_out"], r["actual_repair_hours"],
                          r["target_repair_hours"], r["productivity_ratio"], r["garage_name"],
                          r["cost"], r["status"]])
    
    # 6. Requisitions Sheet
    ws_req = wb.create_sheet("Requisitions")
    req_headers = ["request_no", "requester_name", "place_of_departure", "destination", "purpose",
                  "travel_date", "travel_time", "return_date", "return_time", "num_passengers", "status"]
    ws_req.append(req_headers)
    for cell in ws_req[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for r in requisitions:
        ws_req.append([r["request_no"], r["requester_name"], r["place_of_departure"], r["destination"],
                      r["purpose"], r["travel_date"], r["travel_time"], r["return_date"],
                      r["return_time"], r["num_passengers"], r["status"]])
    
    # Adjust column widths for all sheets
    for ws in [ws_staff, ws_vehicles, ws_fuel, ws_routes, ws_repairs, ws_req]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"✅ Test data saved to: {OUTPUT_FILE}")
    print(f"\n📊 Generated:")
    print(f"  - {len(staff)} staff members")
    print(f"  - {len(vehicles)} vehicles")
    print(f"  - {len(fuel_records)} fuel records")
    print(f"  - {len(routes)} route records")
    print(f"  - {len(repairs)} repair records")
    print(f"  - {len(requisitions)} requisitions")
    
    return OUTPUT_FILE

if __name__ == "__main__":
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.run(["pip", "install", "openpyxl"], check=True)
        from openpyxl import Workbook
    
    output = create_excel_file()
    print(f"\n💾 File ready: {output}")
