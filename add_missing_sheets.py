#!/usr/bin/env python3
"""
Fleet Management System - COMPLETE Test Data Generator
Includes ALL modules: Staff, Fleet, Routes, Fuel, Repairs, Accidents, Requisitions
"""

import random
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

INPUT_FILE = "/root/.openclaw/workspace/fleet-system/fleet_test_data_aligned.xlsx"
OUTPUT_FILE = "/root/.openclaw/workspace/fleet-system/fleet_test_data_complete.xlsx"

# Data pools
FIRST_NAMES = ["John", "Mary", "James", "Patricia", "Robert", "Jennifer", "Michael", "Linda", 
               "William", "Elizabeth", "David", "Susan", "Richard", "Jessica", "Joseph", 
               "Sarah", "Thomas", "Karen", "Charles", "Nancy"]

LAST_NAMES = ["Kamau", "Odhiambo", "Ochieng", "Wanjiku", "Kipchoge", "Mutua", "Omondi", 
              "Wanjiru", "Mwangi", "Kiptoo", "Korir", "Ruto", "Kiprotich", "Tanui",
              "Kiplagat", "Chebet", "Langat", "Kosgei", "Biwott", "Too"]

DEPARTMENTS = ["Transport", "Security", "Operations", "Finance", "HR", "IT", "Sales", "Logistics"]
BRANCHES = ["Nairobi HQ", "Mombasa", "Kisumu", "Nakuru", "Eldoret", "Thika", "Malindi"]

ACCIDENT_TYPES = ["Collision", "Pedestrian", "Property Damage", "Rollover", "Near Miss"]
ACCIDENT_SEVERITY = ["Minor", "Major", "Fatal"]
WEATHER_CONDITIONS = ["Clear", "Rainy", "Foggy", "Overcast"]
ROAD_CONDITIONS = ["Dry", "Wet", "Potholed", "Under Construction", "Muddy"]

ROUTE_ORIGINS = ["Nairobi HQ", "JKIA", "Industrial Area", "Westlands", "Mombasa Road Depot"]
ROUTE_DESTINATIONS = ["Mombasa", "Nairobi", "Kisumu", "Nakuru", "Eldoret", "Malindi"]

def random_date(start_days=365, end_days=0):
    days = random.randint(end_days, start_days)
    return (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

def random_datetime(start_days=365, end_days=0):
    days = random.randint(end_days, start_days)
    hours = random.randint(6, 22)
    minutes = random.randint(0, 59)
    dt = datetime.now() - timedelta(days=days)
    dt = dt.replace(hour=hours, minute=minutes, second=0)
    return dt.strftime('%Y-%m-%d %H:%M')

def random_time():
    return f"{random.randint(6, 18):02d}:{random.randint(0, 59):02d}"

def generate_registration():
    prefixes = ['K', 'K', 'K', 'KB', 'KC', 'KX', 'KY']
    prefix = random.choice(prefixes)
    if len(prefix) == 1:
        letter = random.choice(['A', 'B', 'C', 'D', 'X', 'Y', 'Z'])
        numbers = random.randint(10, 999)
        suffix = random.choice(['A', 'B', 'C', 'X', 'Y'])
        suffix_num = random.randint(1000, 9999)
        return f"{prefix}{letter}{numbers} {suffix}{suffix_num}"
    else:
        numbers = random.randint(100, 999)
        suffix = random.choice(['A', 'B', 'C', 'X', 'Y'])
        suffix_num = random.randint(1000, 9999)
        return f"{prefix}{numbers} {suffix}{suffix_num}"

def generate_accidents(staff_list, vehicle_list, count=10):
    """Generate accident records matching backend import expectations"""
    accidents = []
    drivers = [s for s in staff_list if s.get('role') == 'Driver']
    
    for i in range(count):
        driver = random.choice(drivers) if drivers else None
        vehicle = random.choice(vehicle_list) if vehicle_list else None
        severity = random.choice(ACCIDENT_SEVERITY)
        
        accidents.append({
            "case_number": f"ACC-2025-{1001+i:04d}",
            "accident_date": random_datetime(180, 0),
            "gps_location": f"{random.uniform(-1.5, -1.0):.4f}, {random.uniform(36.7, 37.1):.4f}",
            "registration": vehicle.get('registration', generate_registration()) if vehicle else generate_registration(),
            "driver": driver.get('staff_no', f'ST{1001+i}') if driver else f'ST{1001+i}',
            "accident_type": random.choice(ACCIDENT_TYPES),
            "severity": severity,
            "injuries": random.choice([True, False]),
            "police": True if severity in ["Major", "Fatal"] else random.choice([True, False]),
            "third_party": random.choice([True, False]),
            "weather": random.choice(WEATHER_CONDITIONS),
            "road": random.choice(ROAD_CONDITIONS),
            "description": f"Accident on {random.choice(['highway', 'urban road', 'rural road'])}. " +
                          f"Vehicle {random.choice(['collided with', 'skidded on'])} " +
                          random.choice(["another vehicle", "road barrier", "pothole"]),
            "status": random.choice(["Reported", "Under Investigation", "Closed"])
        })
    return accidents

def generate_requisitions(staff_list, count=15):
    """Generate requisition records matching backend import expectations"""
    requisitions = []
    requesters = [s for s in staff_list if s.get('role') != 'Driver']
    
    for i in range(count):
        requester = random.choice(requesters) if requesters else None
        travel_date = random_date(60, -30)
        
        origin = random.choice(ROUTE_ORIGINS)
        dest = random.choice(ROUTE_DESTINATIONS)
        
        requisitions.append({
            "requested_by": requester.get('staff_no', f'ST{2000+i}') if requester else f'ST{2000+i}',
            "departure": origin,
            "destination": dest,
            "purpose": random.choice(["Client meeting", "Site inspection", "Training", 
                                      "Conference", "Delivery", "Emergency response"]),
            "travel_date": travel_date,
            "travel_time": random_time(),
            "return_date": (datetime.strptime(travel_date, '%Y-%m-%d') + timedelta(days=random.randint(0, 3))).strftime('%Y-%m-%d'),
            "return_time": random_time(),
            "passengers": random.randint(1, 5),
            "passenger_names": requester.get('name', 'Staff Member') if requester else 'Staff Member',
            "status": random.choice(["Draft", "Submitted", "Manager Approved", "Allocated", "Completed"])
        })
    return requisitions

def add_sheets_to_existing():
    print("📂 Loading existing test data...")
    wb = load_workbook(INPUT_FILE)
    
    # Get staff and fleet data from existing sheets for relationships
    staff_data = []
    ws_staff = wb['Staff']
    headers = [cell.value for cell in ws_staff[1]]
    for row in ws_staff.iter_rows(min_row=2, values_only=True):
        if row[0]:
            staff_data.append(dict(zip(headers, row)))
    
    fleet_data = []
    ws_fleet = wb['Fleet']
    headers = [cell.value for cell in ws_fleet[1]]
    for row in ws_fleet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            fleet_data.append(dict(zip(headers, row)))
    
    print(f"   Found {len(staff_data)} staff, {len(fleet_data)} vehicles")
    
    # Generate new data
    accidents = generate_accidents(staff_data, fleet_data, 12)
    requisitions = generate_requisitions(staff_data, 15)
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # === Add Accidents Sheet ===
    if 'Accidents' not in wb.sheetnames:
        ws_acc = wb.create_sheet("Accidents")
        acc_headers = ["case_number", "accident_date", "gps_location", "registration", "driver",
                       "accident_type", "severity", "injuries", "police", "third_party",
                       "weather", "road", "description", "status"]
        ws_acc.append(acc_headers)
        for cell in ws_acc[1]:
            cell.fill = header_fill
            cell.font = header_font
        for a in accidents:
            ws_acc.append([a[h] for h in acc_headers])
        print(f"   ✅ Added Accidents sheet: {len(accidents)} records")
    
    # === Add Requisitions Sheet ===
    if 'Requisitions' not in wb.sheetnames:
        ws_req = wb.create_sheet("Requisitions")
        req_headers = ["requested_by", "departure", "destination", "purpose", 
                       "travel_date", "travel_time", "return_date", "return_time",
                       "passengers", "passenger_names", "status"]
        ws_req.append(req_headers)
        for cell in ws_req[1]:
            cell.fill = header_fill
            cell.font = header_font
        for r in requisitions:
            ws_req.append([r[h] for h in req_headers])
        print(f"   ✅ Added Requisitions sheet: {len(requisitions)} records")
    
    # Auto-adjust columns for new sheets
    for sheet_name in ['Accidents', 'Requisitions']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Complete test data saved: {OUTPUT_FILE}")
    print(f"\n📁 All sheets: {', '.join(wb.sheetnames)}")
    return OUTPUT_FILE

if __name__ == "__main__":
    add_sheets_to_existing()
